import json
import os
import sys
from agents.react_loop import ReactAgent
from tools.calendar import read_calendar
from tools.contact import lookup_contact
from openpyxl import Workbook


def export_outputs_to_excel(
    output_path: str,
    agent_trace: dict | None = None,
    calendar_output: dict | None = None,
    contact_output: dict | None = None,
):
    wb = Workbook()

    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(["Section", "Included"])
    ws_summary.append(["Agent", bool(agent_trace)])
    ws_summary.append(["Calendar", bool(calendar_output)])
    ws_summary.append(["Contact", bool(contact_output)])

    if agent_trace:
        ws_final = wb.create_sheet("AgentFinal")
        final = agent_trace.get("final", {}) or {}
        ws_final.append(["Final Summary"]) 
        for k, v in final.items():
            ws_final.append([k, json.dumps(v, ensure_ascii=False) if isinstance(v, (dict, list)) else v])

        ws_input = wb.create_sheet("AgentInput")
        inp = agent_trace.get("input", {}) or {}
        ws_input.append(["Input"]) 
        for k, v in inp.items():
            ws_input.append([k, json.dumps(v, ensure_ascii=False) if isinstance(v, (dict, list)) else v])

        ws_trace = wb.create_sheet("AgentTrace")
        steps = agent_trace.get("trace", []) or []
        ws_trace.append(["step", "timestamp", "thought", "action", "tool", "action_input", "observation"])
        for s in steps:
            action_input = s.get("action_input", {}) or {}
            obs = s.get("observation", {}) or {}
            tool = None
            if isinstance(action_input, dict):
                tool = action_input.get("tool")
            ws_trace.append([
                s.get("step"),
                s.get("timestamp"),
                s.get("thought"),
                s.get("action"),
                tool,
                json.dumps(action_input, ensure_ascii=False),
                json.dumps(obs, ensure_ascii=False),
            ])

    if calendar_output:
        ws_cal = wb.create_sheet("Calendar")
        slots = calendar_output.get("available_slots", []) or []
        ws_cal.append(["Available Slots"]) 
        ws_cal.append(["slot"]) 
        for s in slots:
            ws_cal.append([s])
        ws_cal.append([])
        events = calendar_output.get("events", []) or []
        if events:
            headers = list(events[0].keys())
            ws_cal.append(headers)
            for e in events:
                ws_cal.append([e.get(h) for h in headers])
        else:
            ws_cal.append(["Events"]) 
            ws_cal.append(["(none)"])

    if contact_output:
        ws_contact = wb.create_sheet("Contact")
        ws_contact.append(["Lookup Result"]) 
        for k, v in (contact_output or {}).items():
            ws_contact.append([k, json.dumps(v, ensure_ascii=False) if isinstance(v, (dict, list)) else v])

    out_dir = os.path.dirname(output_path)
    if out_dir and not os.path.exists(out_dir):
        os.makedirs(out_dir, exist_ok=True)

    wb.save(output_path)


def main():
    # Support running without args by using a friendly default
    if len(sys.argv) < 3:
        subject = "Meeting request"
        body = "Can we schedule for tomorrow?"
        print("No arguments provided. Using default subject/body.")
        print("Usage: python src/main.py \"Subject\" \"Body\"")
    else:
        subject = sys.argv[1]
        body = sys.argv[2]

    agent = ReactAgent(max_steps=6)
    trace = agent.run(subject, body, context={"sender": "manager@company.com"})

    print(json.dumps(trace, indent=2))

    # Optional Excel export (interactive prompt)
    try:
        default_out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "data", "agent_outputs.xlsx")
        choice = input("\nExport outputs to Excel? (y=save to default, c=choose path, N=skip) [y/N/c]: ").strip().lower()
        if choice in ("y", "yes"):
            out_path = default_out
        elif choice in ("c", "choose", "custom"):
            out_path = input(f"Enter Excel output path [{default_out}]: ").strip() or default_out
        else:
            out_path = None

        if out_path:
            # Gather tool outputs for inclusion
            cal = read_calendar(user_id="me", date_hint="next available")
            # Try to lookup a useful contact based on sender
            contact = lookup_contact("manager@company.com")
            export_outputs_to_excel(out_path, agent_trace=trace, calendar_output=cal, contact_output=contact)
            print(f"\nExcel results saved to: {out_path}")
    except Exception as e:
        print("\nExcel export failed:", e)


if __name__ == "__main__":
    main()
